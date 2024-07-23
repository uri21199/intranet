from ..database.conexion import conectar_db, cerrar_db
from openpyxl import Workbook
from io import BytesIO
import mysql.connector
from flask import send_file
# Función para verificar si el usuario existe en la base de datos
def descargar_nomina(personal):
    conexion = conectar_db()

    if conexion.is_connected():
        cursor = conexion.cursor()

        # Crea una cadena de marcadores de posición para los nombres de los empleados en la lista personal
        placeholders = ', '.join(['%s' for _ in range(len(personal))])

        # Construye la consulta SQL con la cláusula WHERE usando los marcadores de posición
        consulta = f"SELECT empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, categoria, convenio, medife FROM nomina WHERE cuenta = 'SI' AND empleado IN ({placeholders}) ORDER BY empleado"

        # Ejecuta la consulta SQL con los nombres de los empleados como argumentos
        cursor.execute(consulta, personal)
        nomina = cursor.fetchall()
        output = generar_excel(nomina)
        cursor.close()
        return send_file(
            output,
            as_attachment=True,
            download_name="nomina.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    cerrar_db(conexion)


    # Función para verificar si el usuario existe en la base de datos
def descargar_nomina_bajas(area):
    conexion = conectar_db()

    if conexion.is_connected():
        cursor = conexion.cursor()
        # Construye la consulta SQL con la cláusula WHERE usando los marcadores de posición
        consulta = "SELECT empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, categoria, convenio, medife FROM nomina WHERE cuenta = 'No' AND area = %s ORDER BY empleado"

        # Ejecuta la consulta SQL con los nombres de los empleados como argumentos
        cursor.execute(consulta, (area,))
        nomina = cursor.fetchall()
        output = generar_excel(nomina)
        cursor.close()
        return send_file(
            output,
            as_attachment=True,
            download_name="nomina.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    cerrar_db(conexion)

def generar_excel(nomina):
    workbook = Workbook()
    sheet = workbook.active

    encabezados = ["Empleado", "Cuil", "Fecha_ingreso", "Finaliza_pp", "Legajo", "Mail", "Cuenta", "Genero", "Fecha_nacimiento", "Forma", "Turno", "Area", "Jerarquia", "Categoria", "Convenio", "Medife"]
    sheet.append(encabezados)

    for fila in nomina:
        sheet.append(fila)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output

def descargar_inventario_recursos():
    conexion = conectar_db()

    if conexion.is_connected():
        cursor = conexion.cursor()

        # Construye la consulta SQL para obtener el inventario de recursos
        consulta = "SELECT * FROM recursos"  # Ajusta la consulta según tu estructura de base de datos

        try:
            cursor.execute(consulta)
            inventario_recursos = cursor.fetchall()

            if inventario_recursos:  # Verifica si hay resultados
                output = generar_excel_inventario(inventario_recursos)  # Función para generar el Excel
                cursor.close()
                return send_file(
                    output,
                    as_attachment=True,
                    download_name="inventario_recursos.xlsx",
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                return "No se encontraron registros en el inventario de recursos."

        except Exception as e:
            print(f"Error al ejecutar la consulta: {e}")
            return "Ocurrió un error al procesar la solicitud."

        finally:
            cerrar_db(conexion)
    
    return "No se pudo establecer conexión con la base de datos."

def generar_excel_inventario(nomina):
    workbook = Workbook()
    sheet = workbook.active

    encabezados = ["ID", "Tipo de equipo", "Estado", "Ubicacion", "Usuario", "Anydesk", "Nro Serie", "Marca", "Modelo", "Ficha", "Area", "Cliente", "Comentarios"]
    sheet.append(encabezados)

    for fila in nomina:
        sheet.append(fila)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


# Función para verificar si el usuario existe en la base de datos
def obtener_dias_empleado(empleado, mes):
    conexion = conectar_db()

    if conexion.is_connected():
        cursor = conexion.cursor()

        # Lista para almacenar los registros filtrados
        novedades = []

        tablas = ['dia_home', 'dia_ausencia', 'dia_estudio', 'dia_vacaciones']

        for tabla in tablas:
            consulta = "SELECT empleado, fecha_inicio, fecha_final, estado, fecha_solicitud, fecha_cambio_estado, superior, area, causa FROM {} WHERE empleado = %s AND (estado = 'SOLICITADO' OR estado = 'APROBADO') AND MONTH(fecha_inicio) = %s".format(tabla)
            try:
                cursor.execute(consulta, (empleado, mes)) 
                resultados = cursor.fetchall()
                for resultado in resultados:
                    novedades.append({'empleado': resultado[0], 'fecha_inicio': resultado[1], 'fecha_final': resultado[2], 'estado': resultado[3], 'concepto': tabla, 'fecha_solicitud': resultado[4], 'fecha_cambio_estado': resultado[5], 'superior': resultado[6], 'area': resultado[7], 'causa': resultado[8]})
            except mysql.connector.Error as err:
                print(f"Error al ejecutar la consulta en la tabla {tabla}: {err}")
        cerrar_db(conexion)
        print(novedades)
        return novedades
    else:
        print("No se pudo realizar la conexión a la base de datos.")
    
    return []



def generar_novedades(novedades):
    workbook = Workbook()
    sheet = workbook.active

    # Encabezados para las nuevas columnas
    encabezados = ["Empleado", "Fecha"]
    sheet.append(encabezados)

    # Agregar filas con los datos de las novedades
    for fila in novedades:
        sheet.append(fila)

    # Guardar el archivo Excel en un objeto BytesIO para enviarlo como respuesta
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output